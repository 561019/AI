from __future__ import annotations

import csv
from datetime import date, datetime, time
from pathlib import Path
from time import perf_counter
from typing import Any

from openpyxl import load_workbook

from framework.core import ROOT, record_interface_call, standard_response
from framework.envelope import make_internal_envelope
from framework.http import post_json
from framework.module_catalog import MODULE_BY_CODE


MODULE = MODULE_BY_CODE["document-table-parsing"]
UPLOAD_ROOT = ROOT / "framework" / "data" / "foundation_data" / "objects" / "uploads"
MAX_EXTRACTED_FIELDS_PER_FILE = 50_000


def get(handler: Any) -> bool:
    if handler.path == "/api/v1/capabilities":
        handler.send(200, {"items": [{"capability_code": item, "enabled": True} for item in MODULE.capabilities]})
        return True
    return False


def post(handler: Any, envelope: dict[str, Any]) -> None:
    if handler.path != MODULE.interface:
        handler.send(404, {"error": {"code": "RESOURCE_NOT_FOUND"}})
        return
    capability = envelope.get("target", {}).get("capability") or envelope.get("action")
    if capability not in MODULE.capabilities:
        handler.send(422, standard_response(envelope, "failed", error={
            "code": "CAPABILITY_NOT_SUPPORTED_BY_MODULE", "capability": capability,
        }))
        return

    payload = envelope.get("payload") if isinstance(envelope.get("payload"), dict) else {}
    documents = payload.get("uploaded_documents") if isinstance(payload.get("uploaded_documents"), list) else []
    if not documents:
        handler.send(422, standard_response(envelope, "failed", error={
            "code": "UPLOADED_DOCUMENTS_REQUIRED",
            "message": "document parsing requires at least one uploaded document reference",
        }))
        return

    started = perf_counter()
    try:
        if capability == "document.package.build":
            result = _build_package(envelope, documents)
        else:
            result = _extract_documents(envelope, documents)
    except ValueError as exc:
        handler.send(422, standard_response(envelope, "failed", error={"code": "DOCUMENT_INPUT_INVALID", "message": str(exc)}))
        return
    except OSError as exc:
        handler.send(500, standard_response(envelope, "failed", error={"code": "DOCUMENT_PARSE_IO_ERROR", "message": str(exc)}))
        return

    record_interface_call(
        trace_id=str(envelope.get("trace_id") or "untraced"),
        source=envelope.get("source") or {"layer": "unknown", "module": "unknown"},
        target={"layer": "business_engine", "module": MODULE.code},
        capability=str(capability), method="POST", url="http://127.0.0.1:8036" + MODULE.interface,
        request={"document_count": len(documents), "file_ids": [_file_id(item) for item in documents]},
        response=result, status_code=200, duration_ms=(perf_counter() - started) * 1000,
    )
    handler.send(200, standard_response(envelope, "success", data=result))


def _build_package(envelope: dict[str, Any], documents: list[Any]) -> dict[str, Any]:
    jobs = [_job_record(envelope, item, state="packaged") for item in documents]
    _persist(envelope, [{"dataset": "parse_jobs", "operation": "upsert", "records": jobs}])
    return {
        "state": "completed", "module": MODULE.code, "platform_capability": "document.package.build",
        "package": {"document_count": len(jobs), "parse_job_ids": [job["parse_job_id"] for job in jobs]},
        "storage": {"datasets": ["parse_jobs"], "writer": "data-operation"},
    }


def _extract_documents(envelope: dict[str, Any], documents: list[Any]) -> dict[str, Any]:
    jobs: list[dict[str, Any]] = []
    fields: list[dict[str, Any]] = []
    summaries: list[dict[str, Any]] = []
    for document in documents:
        job = _job_record(envelope, document, state="running")
        cached = _load_cached_parse_job(envelope, job)
        if cached:
            jobs.append(cached)
            summaries.append({
                "parse_job_id": cached["parse_job_id"],
                "file_id": cached["file_id"],
                "original_name": cached["original_name"],
                "sha256": cached.get("sha256"),
                "state": "completed",
                "parser": cached.get("parser") or "cached",
                "table_count": int(cached.get("table_count") or 0),
                "sheet_names": cached.get("sheet_names") or [],
                "field_count": int(cached.get("field_count") or 0),
                "cache_hit": True,
                "reused_from_parse_job_id": cached.get("reused_from_parse_job_id") or cached["parse_job_id"],
                "evidence_preview": [],
            })
            continue
        file_path = _safe_uploaded_path(document)
        try:
            values, detail = _extract_values(file_path)
            job.update({
                "state": "completed",
                "field_count": len(values),
                "table_count": detail["table_count"],
                "sheet_names": detail.get("sheet_names") or [],
                "parser": detail["parser"],
                "review_required": False,
            })
            for index, value in enumerate(values, start=1):
                fields.append({
                    "record_id": f"{job['parse_job_id']}:field:{index}", "parse_job_id": job["parse_job_id"],
                    "file_id": job["file_id"], "field_name": value["field_name"], "value": value["value"],
                    "value_type": value["value_type"], "source": value["source"],
                    "object_id": job.get("object_id"), "original_name": job.get("original_name"),
                    "sha256": job.get("sha256"),
                    **_ownership(envelope),
                })
            summaries.append({
                "parse_job_id": job["parse_job_id"], "file_id": job["file_id"], "original_name": job["original_name"],
                "sha256": job.get("sha256"), "state": "completed",
                **detail, "field_count": len(values), "evidence_preview": _evidence_preview(values, job["original_name"]),
            })
        except ValueError as exc:
            # Unsupported files are retained as an auditable parse job and explicitly
            # routed to review. They are not reported as an infrastructure outage.
            job.update({"state": "review_required", "field_count": 0, "table_count": 0, "parser": "not_available", "review_required": True, "review_reason": str(exc)})
            summaries.append({"file_id": job["file_id"], "original_name": job["original_name"], "state": "review_required", "reason": str(exc), "field_count": 0, "table_count": 0})
        jobs.append(job)

    writes: list[dict[str, Any]] = [{"dataset": "parse_jobs", "operation": "upsert", "records": jobs}]
    if fields:
        writes.append({"dataset": "extracted_fields", "operation": "upsert", "records": fields})
    _persist(envelope, writes)
    completed = sum(item["state"] == "completed" for item in summaries)
    return {
        "state": "completed" if completed == len(summaries) else "completed_with_review_required",
        "module": MODULE.code, "platform_capability": envelope.get("target", {}).get("capability") or envelope.get("action"),
        "documents": summaries, "field_count": sum(int(item.get("field_count") or 0) for item in summaries),
        "storage": {"datasets": [item["dataset"] for item in writes], "writer": "data-operation"},
    }


def _load_cached_parse_job(envelope: dict[str, Any], job: dict[str, Any]) -> dict[str, Any] | None:
    """Read a prior successful parse through the data-operation engine."""
    cached = _query_cached_parse_job(envelope, {"sha256": job["sha256"], "state": "completed"}) if job.get("sha256") else None
    if not cached:
        cached = _query_cached_parse_job(envelope, {"file_id": job["file_id"], "state": "completed"})
    if not isinstance(cached, dict):
        return None
    if job.get("sha256") and cached.get("sha256") != job.get("sha256"):
        return None
    parse_job_id = str(cached.get("parse_job_id") or job["parse_job_id"])
    referenced_file_ids = _unique_strings([
        *(
            cached.get("referenced_file_ids")
            if isinstance(cached.get("referenced_file_ids"), list)
            else [cached.get("file_id")]
        ),
        job.get("file_id"),
    ])
    return {
        **cached,
        **_ownership(envelope),
        "parse_job_id": parse_job_id,
        "record_id": str(cached.get("record_id") or parse_job_id),
        "file_id": job["file_id"],
        "object_id": job.get("object_id") or cached.get("object_id"),
        "original_name": job["original_name"],
        "sha256": job.get("sha256") or cached.get("sha256"),
        "referenced_file_ids": referenced_file_ids,
        "reused_from_parse_job_id": parse_job_id,
    }


def _query_cached_parse_job(envelope: dict[str, Any], filters: dict[str, Any]) -> dict[str, Any] | None:
    task_id = str((envelope.get("payload") or {}).get("platform_task_id") or envelope.get("request_id"))
    inner = make_internal_envelope(
        str(envelope.get("trace_id")),
        envelope.get("actor") or {},
        task_id,
        "data.search",
        "business_engine",
        "engine-gateway",
        {
            "dataset": "parse_jobs",
            "filters": filters,
            "limit": 1,
            "compact": True,
        },
        source_layer="business_engine",
        source_module=MODULE.code,
        context=envelope.get("context") if isinstance(envelope.get("context"), dict) else None,
    )
    try:
        status, response = post_json(
            "http://127.0.0.1:8200/api/v1/engine/instructions",
            inner,
            timeout=10,
            caller={"layer": "business_engine", "module": MODULE.code},
        )
    except OSError:
        return None
    if status not in {200, 202} or not isinstance(response, dict) or response.get("status") != "success":
        return None
    data = response.get("data") or {}
    storage = data.get("storage_result") if isinstance(data, dict) else {}
    items = storage.get("items") if isinstance(storage, dict) else []
    if not isinstance(items, list) or not items:
        return None
    cached = items[0]
    return cached if isinstance(cached, dict) else None


def _extract_values(path: Path) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    suffix = path.suffix.lower()
    if suffix == ".xlsx":
        return _extract_xlsx(path)
    if suffix in {".csv", ".txt"}:
        return _extract_csv(path)
    raise ValueError(f"unsupported document type {suffix or '(no extension)'}; supported types are .xlsx, .csv and .txt")


def _extract_xlsx(path: Path) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    values: list[dict[str, Any]] = []
    try:
        for sheet in workbook.worksheets:
            rows = list(sheet.iter_rows(values_only=True))
            header_index = _header_row_index(rows)
            if header_index is None:
                continue
            headers = [str(cell).strip() if cell not in (None, "") else f"column_{index + 1}" for index, cell in enumerate(rows[header_index])]
            record_key_index = next((index for index, header in enumerate(headers) if header.lower() in {"contract_id", "record_id", "id"}), None)
            for row_number, row in enumerate(rows[header_index + 1:], start=header_index + 2):
                record_key = row[record_key_index] if record_key_index is not None and record_key_index < len(row) else None
                for column_number, raw_value in enumerate(row, start=1):
                    if raw_value in (None, ""):
                        continue
                    if len(values) >= MAX_EXTRACTED_FIELDS_PER_FILE:
                        raise ValueError(f"document exceeds {MAX_EXTRACTED_FIELDS_PER_FILE} non-empty cells")
                    values.append(_value(headers[column_number - 1], raw_value, {
                        "sheet": sheet.title, "cell": f"R{row_number}C{column_number}", "row": row_number,
                        "column": column_number, "record_key": str(record_key) if record_key not in (None, "") else None,
                    }))
    finally:
        workbook.close()
    return values, {"parser": "openpyxl", "table_count": len(workbook.sheetnames), "sheet_names": workbook.sheetnames}


def _extract_csv(path: Path) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    with path.open("r", encoding="utf-8-sig", errors="replace", newline="") as handle:
        rows = list(csv.reader(handle))
    if not rows:
        return [], {"parser": "csv", "table_count": 0, "sheet_names": []}
    headers = [cell.strip() or f"column_{index + 1}" for index, cell in enumerate(rows[0])]
    values: list[dict[str, Any]] = []
    for row_number, row in enumerate(rows[1:], start=2):
        for column_number, raw_value in enumerate(row, start=1):
            if not raw_value.strip():
                continue
            if len(values) >= MAX_EXTRACTED_FIELDS_PER_FILE:
                raise ValueError(f"document exceeds {MAX_EXTRACTED_FIELDS_PER_FILE} non-empty cells")
            values.append(_value(headers[column_number - 1], raw_value, {"sheet": "csv", "cell": f"R{row_number}C{column_number}", "row": row_number, "column": column_number}))
    return values, {"parser": "csv", "table_count": 1, "sheet_names": ["csv"]}


def _header_row_index(rows: list[tuple[Any, ...]]) -> int | None:
    """Prefer the field-name row over a merged title row at the top of a workbook."""
    best_index: int | None = None
    best_score = 0
    for index, row in enumerate(rows[:30]):
        values = [str(cell).strip() for cell in row if cell not in (None, "")]
        if not values:
            continue
        unique = len(set(values))
        field_like = sum("_" in value or value.lower().endswith(("id", "amount", "status", "date")) for value in values)
        score = unique * 10 + field_like * 20
        if score > best_score:
            best_index, best_score = index, score
    return best_index


def _evidence_preview(values: list[dict[str, Any]], original_name: str) -> list[dict[str, Any]]:
    wanted = {
        "contract_id", "tail_payment_due", "received_amount", "risk_note", "invoice_id",
        "file_name", "upload_status", "attachment_ref", "contract_status",
    }
    preview = []
    for item in values:
        if str(item.get("field_name") or "").lower() not in wanted:
            continue
        preview.append({
            "file_name": original_name, "field_name": item["field_name"], "value": item["value"],
            "source": item["source"],
        })
    return preview[:120]


def _value(field_name: str, raw_value: Any, source: dict[str, Any]) -> dict[str, Any]:
    if isinstance(raw_value, (datetime, date, time)):
        raw_value = raw_value.isoformat()
        value_type = "datetime"
    elif isinstance(raw_value, bool):
        value_type = "boolean"
    elif isinstance(raw_value, (int, float)):
        value_type = "number"
    else:
        value_type = "text"
    return {"field_name": field_name, "value": raw_value, "value_type": value_type, "source": source}


def _job_record(envelope: dict[str, Any], document: Any, *, state: str) -> dict[str, Any]:
    if not isinstance(document, dict):
        raise ValueError("each uploaded document must be an object")
    file_id = _file_id(document)
    if not file_id:
        raise ValueError("uploaded document file_id is required")
    sha256 = document.get("sha256")
    parse_key = f"sha256-{str(sha256)[:32]}" if sha256 else file_id
    return {
        # A parse result belongs to the immutable uploaded object, not to one
        # workflow run. Stable identifiers make later writes idempotent.
        "parse_job_id": f"parse-{parse_key}",
        "record_id": f"parse-{parse_key}",
        "file_id": file_id, "object_id": document.get("object_id"),
        "original_name": document.get("original_name") or document.get("name") or file_id,
        "sha256": sha256, "state": state, "referenced_file_ids": [file_id], **_ownership(envelope),
    }


def _file_id(document: Any) -> str:
    return str(document.get("file_id") or "") if isinstance(document, dict) else ""


def _ownership(envelope: dict[str, Any]) -> dict[str, Any]:
    actor = envelope.get("actor") or {}
    context = envelope.get("context") or {}
    return {
        "tenant_id": actor.get("tenant_id"), "owner_account_id": actor.get("user_id") or actor.get("actor_id"),
        "project_id": context.get("project_id"), "conversation_id": context.get("conversation_id"),
    }


def _unique_strings(values: list[Any]) -> list[str]:
    result: list[str] = []
    for value in values:
        if value in (None, ""):
            continue
        text = str(value)
        if text not in result:
            result.append(text)
    return result


def _safe_uploaded_path(document: Any) -> Path:
    raw_path = str(document.get("saved_path") or "")
    if not raw_path:
        uri = str(document.get("storage_uri") or "")
        raw_path = str(UPLOAD_ROOT / Path(uri).name) if uri else ""
    path = Path(raw_path).resolve()
    try:
        path.relative_to(UPLOAD_ROOT.resolve())
    except ValueError as exc:
        raise ValueError("uploaded document path is outside the managed upload store") from exc
    if not path.is_file():
        raise ValueError("uploaded document object does not exist")
    return path


def _persist(envelope: dict[str, Any], writes: list[dict[str, Any]]) -> None:
    task_id = str((envelope.get("payload") or {}).get("platform_task_id") or envelope.get("request_id"))
    inner = make_internal_envelope(
        str(envelope.get("trace_id")), envelope.get("actor") or {}, task_id,
        "data.persist", "business_engine", "engine-gateway", {"writes": writes},
        source_layer="business_engine", source_module=MODULE.code,
        context=envelope.get("context") if isinstance(envelope.get("context"), dict) else None,
    )
    status, response = post_json(
        "http://127.0.0.1:8200/api/v1/engine/instructions", inner, timeout=30,
        caller={"layer": "business_engine", "module": MODULE.code},
    )
    if status not in {200, 202} or not isinstance(response, dict) or response.get("status") != "success":
        raise OSError(f"parse result persistence failed: {response}")
